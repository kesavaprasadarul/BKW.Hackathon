"""IO"""

from __future__ import annotations
from pathlib import Path
from typing import Dict, Optional, Tuple, Iterable
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.roomtypes.matching import norm_key

HeaderInfo = Tuple[int, Optional[int], Optional[int]]


def load_wb(path: Path):
    """
    data_only=False preserves formulas; keep_vba is False by default.
    """
    return load_workbook(filename=path, data_only=False)


def save_wb(wb, out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def detect_header_xlsx(ws: Worksheet, max_scan_rows: int) -> HeaderInfo:
    """
    Find the header row (1-based), and the columns for:
      - Raum-Bezeichnung (bez_col)
      - Nummer Raumtyp (nr_col)
    """
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        m: Dict[str, int] = {}
        for c_idx, v in enumerate(row_vals, start=1):
            nk = norm_key(v)
            if nk:
                m[nk] = c_idx

        bez_aliases = {
            "raumbezeichnung",
            "raumbezeich",
            "raumbez",
            "raumbezeichng",
            "raumbezchung",
            "raum-bezeichnung",
        }
        nr_aliases = {"nummerraumtyp", "nummer raumtyp"}

        bez_col = next((m[k] for k in bez_aliases if k in m), None)
        nr_col = next((m[k] for k in nr_aliases if k in m), None)
        if bez_col is not None or nr_col is not None:
            return r, bez_col, nr_col

    return None, None, None


def detect_header_mapping(ws: Worksheet, max_scan_rows: int) -> HeaderInfo:
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        m: Dict[str, int] = {}
        for c_idx, v in enumerate(row_vals, start=1):
            nk = norm_key(v)
            if nk:
                m[nk] = c_idx
        nr_aliases = {"nr", "nummer"}
        roomtype_aliases = {"bezeichnung", "raumbezeichnung", "raum-bezeichnung"}

        nr_col = next((m[k] for k in nr_aliases if k in m), None)
        roomtype_col = next((m[k] for k in roomtype_aliases if k in m), None)
        if nr_col is not None or roomtype_col is not None:
            return r, nr_col, roomtype_col
    return None, None, None


def ensure_nr_column(ws: Worksheet, header_row: int, nr_col: Optional[int]) -> int:
    """
    If "Nummer Raumtyp" column is missing, create it as the last column
    """
    if nr_col is not None:
        return nr_col
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = "Nummer Raumtyp"
    return new_col


def iter_data_rows(ws: Worksheet, header_row: int) -> Iterable[int]:
    """
    Yield row indices (1-based) of data rows: header_row+1 .. max_row.
    """
    for r in range(header_row + 1, ws.max_row + 1):
        yield r
